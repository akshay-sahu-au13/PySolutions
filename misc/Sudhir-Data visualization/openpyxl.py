import openpyxl
# you could create a template spreadsheet that does all of the
# worksheet.write() commands on lines 188-195
wb = openpyxl.load_workbook('Template.xlsx')
wb.create_sheet('my_sheet')
ws = wb.get_sheet_by_name(title='my_sheet')
row=1
for i in range(len(titles)):
    ws.cell(row=row, column=0, value=titles[i])
    ws.cell(row=row, column=1, value=Genre_Arr[i])
    ws.cell(row=row, column=2, value=Budget_Arr[i])
    ws.cell(row=row, column=3, value=Domestic_Gross_Arr[i])
    ws.cell(row=row, column=4, value=Release_Date_Arr[i])
    ws.cell(row=row, column=5, value=Theaters_Arr[i])
    ws.cell(row=row, column=6, value=Views_Arr[i])
    ws.cell(row=row, column=7, value=Edits_Arr[i])
    ws.cell(row=row, column=8, value=Editors_Arr[i])
    row += 1
wb.save('my_file.xlsx')